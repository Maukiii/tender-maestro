from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

CV_SECTION_KEYS = [
    "Professional Summary",
    "Education",
    "Professional Experience",
    "Technical Skills",
    "Selected Publications",
    "Selected Publications & Contributions",
    "Selected Project Contributions",
    "Languages",
    "Role in Proposed Project",
]


def _join_chunks(entry: dict[str, Any]) -> str:
    if entry.get("raw_text"):
        return str(entry["raw_text"])
    return "\n".join(entry.get("chunks", []))


def _extract_email(text: str) -> str:
    match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text)
    return match.group(0) if match else ""


def _extract_phone(text: str) -> str:
    match = re.search(r"\+\d[\d\s\-]{7,}\d", text)
    return match.group(0).strip() if match else ""


def _extract_company_name(text: str) -> str:
    match = re.search(r"([A-Z][A-Za-z0-9&.,\-\s]+GmbH)", text)
    return match.group(1).strip() if match else ""


def _extract_location(text: str) -> str:
    match = re.search(r"\b(Berlin|Brussels|Paris|Athens|Seville|Germany|EU)\b", text)
    return match.group(1) if match else ""


def _extract_between(text: str, start: str, end_markers: list[str]) -> str:
    lower = text.lower()
    start_idx = lower.find(start.lower())
    if start_idx == -1:
        return ""
    end_idx = len(text)
    for marker in end_markers:
        idx = lower.find(marker.lower(), start_idx + len(start))
        if idx != -1 and idx < end_idx:
            end_idx = idx
    return text[start_idx:end_idx].strip()


def _extract_past_projects(text: str) -> list[str]:
    projects: list[str] = []
    # Rows from xlsx parser are converted to "col1 | col2 | ..."
    for line in text.splitlines():
        if "|" not in line:
            continue
        cols = [c.strip() for c in line.split("|")]
        if len(cols) < 2:
            continue
        ref = cols[0]
        title = cols[1]
        if "/" in ref and len(title) > 6 and "project title" not in title.lower():
            projects.append(title)
    # Deduplicate but preserve order
    seen: set[str] = set()
    unique_projects: list[str] = []
    for project in projects:
        if project not in seen:
            seen.add(project)
            unique_projects.append(project)
    return unique_projects


def _extract_sheet_block(text: str, sheet_name: str) -> str:
    marker = f"[{sheet_name}]"
    start = text.find(marker)
    if start == -1:
        return ""
    start += len(marker)
    end = text.find("\n[", start)
    if end == -1:
        end = len(text)
    return text[start:end].strip()


def _parse_turnover(text: str) -> dict[str, int]:
    result: dict[str, int] = {}
    for value, year in re.findall(r"([0-9]+(?:\.[0-9]+)?)M\s*\((2024|2023|2022)\)", text):
        result[year] = int(float(value) * 1_000_000)
    return result


def _parse_basic_info(profile_text: str, capabilities_text: str) -> dict[str, Any]:
    combined = f"{profile_text}\n{capabilities_text}"
    legal_name = _extract_company_name(combined)
    legal_name = legal_name.replace("Who We Are", "").strip()
    founded_match = re.search(r"Since\s+(\d{4})", combined, flags=re.IGNORECASE)
    vat_match = re.search(r"\bDE[0-9]{9}\b", combined)
    employees_match = re.search(r"Team size\s*~?(\d+)\s*FTE", combined, flags=re.IGNORECASE)
    website = ""
    email = _extract_email(combined)
    if email and "@" in email:
        website = email.split("@", 1)[1]
    iso_match = re.search(r"(ISO\s*9001[^\n]*)", combined, flags=re.IGNORECASE)
    return {
        "legal_name": legal_name,
        "founded": int(founded_match.group(1)) if founded_match else None,
        "headquarters": "Berlin, Germany",
        "legal_form": "GmbH" if "GmbH" in legal_name else "",
        "VAT_number": vat_match.group(0) if vat_match else "",
        "employees": int(employees_match.group(1)) if employees_match else None,
        "annual_turnover": _parse_turnover(combined),
        "website": website,
        "ISO_certifications": [iso_match.group(0).strip()] if iso_match else [],
        "GDPR_compliance": True,
    }


def _load_sheet_rows(xlsx_path: str, sheet_name: str) -> list[list[str]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    rows: list[list[str]] = []
    for row in wb[sheet_name].iter_rows(values_only=True):
        values = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if values:
            rows.append(values)
    return rows


def _parse_services(profile_text: str) -> list[str]:
    services: list[str] = []
    candidates = [
        "Market and Ecosystem Mapping",
        "Continuous Market Monitoring",
        "Regulatory and Policy Research",
        "Supply Chain and Network Analysis",
    ]
    low = profile_text.lower()
    for item in candidates:
        if item.lower() in low:
            services.append(item)
    return services


def _parse_capabilities(capabilities_text: str) -> dict[str, Any]:
    capability_names = [
        "Market & ecosystem mapping",
        "Evidence-based classification",
        "Continuous monitoring",
        "Supply chain mapping",
        "Regulatory scoping",
        "API delivery",
    ]
    capabilities_list: list[dict[str, str]] = []
    for name in capability_names:
        if name.lower() in capabilities_text.lower():
            capabilities_list.append({"name": name, "description": ""})

    def _stack_between(start: str, end: str) -> list[str]:
        pattern = rf"{re.escape(start)}\s+(.+?)\s+{re.escape(end)}"
        match = re.search(pattern, capabilities_text, flags=re.IGNORECASE | re.DOTALL)
        if not match:
            return []
        raw = match.group(1)
        return [p.strip() for p in raw.split(",") if p.strip()]

    technology_stack = {
        "data_collection": _stack_between("Data collection", "NLP & classification"),
        "nlp_classification": _stack_between("NLP & classification", "Entity resolution"),
        "orchestration": _stack_between("Pipeline orchestration", "Storage & query"),
        "storage": _stack_between("Storage & query", "Delivery"),
        "delivery": _stack_between("Delivery", "Selected Project References"),
    }
    return {
        "capabilities_list": capabilities_list,
        "methodology": "WebMap five-stage pipeline" if "WebMap" in capabilities_text else "",
        "technology_stack": technology_stack,
    }


def _parse_projects(credentials_text: str, credentials_path: str) -> list[dict[str, Any]]:
    rows = _load_sheet_rows(credentials_path, "Project Credentials")
    projects: list[dict[str, Any]] = []
    for row in rows:
        if len(row) < 8:
            continue
        if "Reference / Client" in row[0] or "MERIDIAN" in row[0]:
            continue
        if "/" not in row[0]:
            continue

        leads = re.findall(r"(Dr\.\s*[A-Z]\.\s*[A-Za-z]+|[A-Z]\.\s*[A-Za-z]+)", row[5])
        value_raw = re.sub(r"[^0-9]", "", row[2])
        deliverables = re.split(r"[;\n]", row[6])
        ref_parts = [p.strip() for p in row[0].splitlines() if p.strip()]
        client_name = row[0]
        if len(ref_parts) > 1:
            client_name = ref_parts[1].split(",")[0].strip()
        elif "," in row[0]:
            client_name = row[0].split(",", 1)[0].strip()
        projects.append(
            {
                "project_name": row[1],
                "client": client_name,
                "value_EUR": int(value_raw) if value_raw else None,
                "duration": row[3],
                "role": row[4],
                "lead_researchers": leads if leads else [row[5]] if row[5] else [],
                "key_deliverables": [
                    re.sub(r"^[^\w]+", "", d).strip()
                    for d in deliverables
                    if d.strip()
                ],
                "status": row[7],
            }
        )
    if projects:
        return projects
    # fallback for text-only parsing
    block = _extract_sheet_block(credentials_text, "Project Credentials")
    for line in block.splitlines():
        if "|" in line and "/" in line:
            cols = [c.strip() for c in line.split("|")]
            if len(cols) >= 2:
                projects.append(
                    {
                        "project_name": cols[1],
                        "client": cols[0].split()[0],
                        "value_EUR": None,
                        "duration": cols[3] if len(cols) > 3 else "",
                        "role": cols[4] if len(cols) > 4 else "",
                        "lead_researchers": [],
                        "key_deliverables": [],
                        "status": cols[7] if len(cols) > 7 else "",
                    }
                )
    return projects


def _parse_team_members(credentials_text: str, credentials_path: str) -> list[dict[str, Any]]:
    rows = _load_sheet_rows(credentials_path, "Team Skills Matrix")
    members: list[dict[str, Any]] = []
    for cols in rows:
        if len(cols) < 9:
            continue
        if cols[0].lower() == "name" or "MERIDIAN INTELLIGENCE" in cols[0]:
            continue
        availability_raw = re.sub(r"[^0-9]", "", cols[8])
        members.append(
            {
                "name": cols[0],
                "role": cols[1],
                "seniority": cols[2],
                "skills": {
                    "Market Mapping": cols[3],
                    "Data Engineering": cols[4],
                    "NLP/ML": cols[5],
                    "Policy Analysis": cols[6],
                    "Stakeholder Management": cols[7],
                },
                "availability_days_per_month": int(availability_raw) if availability_raw else None,
            }
        )
    if members:
        return members
    block = _extract_sheet_block(credentials_text, "Team Skills Matrix")
    for line in block.splitlines():
        if "|" not in line:
            continue
        cols = [c.strip() for c in line.split("|")]
        if len(cols) >= 3 and cols[0].lower() != "name":
            members.append(
                {
                    "name": cols[0],
                    "role": cols[1] if len(cols) > 1 else "",
                    "seniority": cols[2] if len(cols) > 2 else "",
                    "skills": {},
                    "availability_days_per_month": None,
                }
            )
    return members


def _parse_reference_contacts(credentials_text: str, credentials_path: str) -> list[dict[str, Any]]:
    rows = _load_sheet_rows(credentials_path, "Reference Contacts")
    refs: list[dict[str, Any]] = []
    for cols in rows:
        if len(cols) < 6:
            continue
        if cols[0].lower() == "project reference" or "REFERENCE CONTACTS" in cols[0]:
            continue
        refs.append(
            {
                "project_reference": cols[0],
                "client_organisation": cols[1],
                "contact_name": cols[2],
                "contact_title": cols[3],
                "email": cols[4],
                "relationship": cols[5],
            }
        )
    if refs:
        return refs
    block = _extract_sheet_block(credentials_text, "Reference Contacts")
    for line in block.splitlines():
        if "|" not in line:
            continue
        cols = [c.strip() for c in line.split("|")]
        if len(cols) >= 6:
            refs.append(
                {
                    "project_reference": cols[0],
                    "client_organisation": cols[1],
                    "contact_name": cols[2],
                    "contact_title": cols[3],
                    "email": cols[4],
                    "relationship": cols[5],
                }
            )
    return refs


def _extract_method_steps(text: str) -> list[str]:
    steps: list[str] = []
    for match in re.finditer(r"(Stage\s+\d+\s*[-:]\s*[^\n.]+)", text, flags=re.IGNORECASE):
        steps.append(match.group(1).strip())
    if steps:
        return steps
    for match in re.finditer(r"(\d+\.\s+[A-Z][^\n]+)", text):
        candidate = match.group(1).strip()
        if len(candidate) > 6:
            steps.append(candidate)
    return steps[:12]


def _extract_tools_used(text: str) -> list[str]:
    known_tools = [
        "Python",
        "Scrapy",
        "Playwright",
        "boto3",
        "spaCy",
        "scikit-learn",
        "sentence-transformers",
        "BERT",
        "Apache Airflow",
        "PostgreSQL",
        "Elasticsearch",
        "S3",
        "AWS",
        "JSON",
        "CSV",
        "Parquet",
        "REST API",
    ]
    found: list[str] = []
    lower = text.lower()
    for tool in known_tools:
        if tool.lower() in lower:
            found.append(tool)
    return found


def _cv_sections(text: str) -> dict[str, Any]:
    result: dict[str, Any] = {key: [] for key in CV_SECTION_KEYS}
    result["Role in Proposed Project"] = ""

    def _clean_markdown(line: str) -> str:
        line = re.sub(r"\*\*(.*?)\*\*", r"\1", line)
        line = re.sub(r"\*(.*?)\*", r"\1", line)
        line = line.replace("`", "")
        line = re.sub(r"\s+", " ", line).strip()
        return line

    def _content_to_items(content: str) -> list[str]:
        items: list[str] = []
        for raw_line in content.splitlines():
            line = raw_line.strip()
            if not line or line == "---":
                continue
            if line.startswith("|---"):
                continue
            if line.startswith("|") and line.endswith("|"):
                cols = [c.strip() for c in line.strip("|").split("|")]
                if len(cols) >= 3 and cols[0].lower() != "project":
                    items.append(f"{cols[0]} - {cols[1]}: {cols[2]}")
                    continue
            line = re.sub(r"^[-*]\s+", "", line)
            line = re.sub(r"^###\s+", "", line)
            line = _clean_markdown(line)
            if line:
                items.append(line)
        return items

    pattern = re.compile(r"^##\s+(.+?)\s*$", flags=re.MULTILINE)
    matches = list(pattern.finditer(text))
    for i, match in enumerate(matches):
        section_name = match.group(1).strip()
        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        content = text[start:end].strip()
        lines = _content_to_items(content)

        if section_name == "Publication":
            section_name = "Selected Publications"

        if section_name == "Role in Proposed Project":
            result["Role in Proposed Project"] = " ".join(lines).strip()
            continue

        if section_name in result:
            if section_name == "Languages":
                expanded: list[str] = []
                for line in lines:
                    for part in line.split(","):
                        cleaned = part.strip()
                        if cleaned:
                            expanded.append(cleaned)
                lines = expanded
            result[section_name] = lines

    return result


def _extract_cv_name_and_title(text: str, fallback_stem: str) -> tuple[str, str]:
    name_match = re.search(r"^#\s+([^\n]+)", text, flags=re.MULTILINE)
    if name_match:
        name = name_match.group(1).strip()
    else:
        first_line = text.splitlines()[0].strip() if text.splitlines() else ""
        name = first_line if first_line else fallback_stem.replace("_", " ").title()
    title_match = re.search(r"\*\*([^*]+)\*\*", text)
    title = title_match.group(1).strip() if title_match else ""
    return name, title


def _build_company_documents(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    files = {Path(item["source"]).name.lower(): item for item in entries if item["category"] == "company_info"}
    profile_text = _join_chunks(files.get("company_profile.docx", {})) if files.get("company_profile.docx") else ""
    capabilities_text = _join_chunks(files.get("capabilities_overview.pdf", {})) if files.get("capabilities_overview.pdf") else ""
    credentials_text = _join_chunks(files.get("credentials_projects.xlsx", {})) if files.get("credentials_projects.xlsx") else ""
    credentials_source = files.get("credentials_projects.xlsx", {}).get("source", "")

    return [
        {
            "basic_info": _parse_basic_info(profile_text, capabilities_text),
            "company_profile": {
                "company_profile_text": profile_text,
                "services": _parse_services(profile_text),
                "approach": "Evidence-based organisation profiling using observable, verifiable signals",
            },
            "capabilities": _parse_capabilities(capabilities_text),
            "projects": _parse_projects(credentials_text, credentials_source),
            "team_members": _parse_team_members(credentials_text, credentials_source),
            "reference_contacts": _parse_reference_contacts(credentials_text, credentials_source),
            "key_facts": {
                "CPV_codes": re.findall(r"\b\d{8}\b", capabilities_text),
                "NUTS_code": "",
                "NACE_code": "",
                "authorized_signatory": "Dr. Anna Becker" if "Dr. Anna Becker" in capabilities_text else "",
                "bank_details": "",
            },
        }
    ]


def _build_methodology_documents(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    docs: list[dict[str, Any]] = []
    for item in entries:
        if item["category"] != "methodology":
            continue
        file_name = Path(item["source"]).name
        text = _join_chunks(item)
        description = _extract_between(
            text,
            "Overview",
            ["Pipeline Architecture", "Seed Universe Construction", "Evidence Classification"],
        )
        if not description:
            description = " ".join(text.splitlines()[:5]).strip()
        docs.append(
            {
                "file_name": file_name,
                "methodology_name": Path(file_name).stem.replace("_", " ").title(),
                "description": description,
                "steps": _extract_method_steps(text),
                "tools_used": _extract_tools_used(text),
            }
        )
    return docs


def _build_team_cvs(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cvs: list[dict[str, Any]] = []
    for item in entries:
        if item["category"] != "team_cv":
            continue
        source = Path(item["source"])
        text = _join_chunks(item)
        name, title = _extract_cv_name_and_title(text, source.stem)
        cvs.append(
            {
                "id": source.stem.replace("cv_", ""),
                "name": name,
                "title": title,
                "contact": {
                    "email": _extract_email(text),
                    "phone": _extract_phone(text),
                    "company": _extract_company_name(text),
                    "location": _extract_location(text),
                },
                "category": "team_cvs",
                "sections": _cv_sections(text),
            }
        )
    return cvs


def build_standardized_output(entries: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "company_documents": _build_company_documents(entries),
        "methodology_documents": _build_methodology_documents(entries),
        "team_cvs": _build_team_cvs(entries),
    }

